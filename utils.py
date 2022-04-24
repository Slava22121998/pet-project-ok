import calendar
import datetime

import openpyxl
from openpyxl.styles import Alignment, Font


def get_data_about_employees(fio: list):
    days_count = calendar.mdays[datetime.date.today().month]  # Получаем количество дней в месяце
    employees_data_dict = dict()
    days_of_duty_list = list()
    start_day = 1
    for item in fio:
        for day in range(start_day, days_count + 1, 3):
            days_of_duty_list.append(str(day))
        if '/' in item:  # Если в смену заступает 2 контролёра
            for name in item.split('/'):
                employees_data_dict[name] = days_of_duty_list
        else:
            employees_data_dict[item] = days_of_duty_list  # Создаем словарь с днями дежурств - нормализованный график
        days_of_duty_list = []
        start_day += 1

    return employees_data_dict


def get_fio_of_employees(info_str: str):  # Функция, которая возвращает список ФИО сотрудников - РАБОТА БД
    info_list = info_str.split(',')
    return info_list


def set_data_of_employees_in_report_card(fio_dct: dict, unit):
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.merge_cells(start_row=1, start_column=6, end_row=2, end_column=10)
    sheet.column_dimensions['B'].width = 30
    sheet.cell(row=1, column=6).value = f'Подразделение ОК СЭБ: {unit}'
    start_col = 4
    start_row = 3
    days_count = calendar.mdays[datetime.date.today().month]
    for day in range(1, days_count + 1):
        sheet.cell(row=start_row, column=start_col).value = day
        sheet.cell(row=start_row, column=start_col).alignment = Alignment(horizontal="center", vertical="center")
        sheet.cell(row=start_row, column=start_col).font = Font(bold=True, color='ff0000', name='Arial', size=12)
        start_col += 1
    start_row = 5
    for i in range(1, len(fio_dct.keys()) + 1):
        sheet.cell(row=start_row, column=1).value = i
        sheet.cell(row=start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        start_row += 1
    start_row = 5
    for fio, days_list in fio_dct.items():
        sheet.cell(row=start_row, column=2).value = fio
        sheet.cell(row=start_row, column=2).alignment = Alignment(horizontal="center", vertical="center")
        for day in days_list:
            if '-' in day:
                sheet.cell(row=start_row, column=int(day.split('-')[0]) + 3).value = int(day.split('-')[1])
                sheet.cell(row=start_row, column=int(day.split('-')[0]) + 3).alignment = Alignment(
                    horizontal="center",
                    vertical="center")
            else:
                sheet.cell(row=start_row, column=int(day) + 3).value = 22
                sheet.cell(row=start_row, column=int(day) + 3).alignment = Alignment(horizontal="center",
                                                                                     vertical="center")
        start_row += 1

    book.save(f'static/excel_files/{unit}_табель.xls')
    book.close()
